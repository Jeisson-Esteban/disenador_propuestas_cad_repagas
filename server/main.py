"""
FastAPI server para RepagasConcept — genera propuestas de cocina industrial a partir del
formulario de cliente, ejecutando el pipeline LLM -> DB -> DXF -> PDFs y devolviendo un ZIP.

Arranque:
    cd server && uvicorn main:app --reload --port 8000

Endpoints:
    POST /generar            — JSON body con FormularioCliente
    POST /generar-con-plano  — multipart/form-data: JSON + archivo DWG/DXF del cliente
    POST /feedback           — aplica cambios del usuario sobre la ultima propuesta generada
    GET  /catalogo           — devuelve equipos activos de la BD agrupados por zona
"""

import io
import json
import os
import shutil
import tempfile
import uuid
import zipfile
from pathlib import Path

from fastapi import FastAPI, File, Form, UploadFile, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse

from core.database import get_db_connection
from features.propuestas.schemas import FormularioCliente
from features.propuestas.llm import generar_propuesta_llm
from features.propuestas.resolver import resolver_equipos
from features.planos.generar import (
    generar_plano,
    generar_plano_integrado,
    imprimir_resumen,
)
from features.documentos.pdf import generar_pdf_prospeccion, generar_pdf_presupuesto
from features.planos.conversion import dwg_a_dxf, _find_oda, _find_libredwg

TIPO_TO_ZONA = {
    "cocina_gas": "coccion", "cocina_electrica": "coccion", "cocina_induccion": "coccion",
    "fry_top_gas": "coccion", "fry_top_electrico": "coccion",
    "freidora_gas": "coccion", "freidora_electrica": "coccion",
    "plancha": "coccion", "barbacoa": "coccion", "marmita": "coccion",
    "bano_maria": "coccion", "cuece_pastas": "coccion", "neutro": "coccion",
    "mantenedor_fritos": "coccion", "soporte": "coccion",
    "mesa_refrig_conservacion": "refrigeracion", "mesa_refrig_congelacion": "refrigeracion",
    "mesa_refrig_conservacion_gn": "refrigeracion", "mesa_refrig_congelacion_gn": "refrigeracion",
    "armario_conservacion": "refrigeracion", "armario_congelacion": "refrigeracion",
    "armario_snack": "refrigeracion", "frente_mostrador": "refrigeracion",
    "mesa_pizza": "refrigeracion", "mesa_ensalada": "refrigeracion",
    "mesa_trabajo": "refrigeracion", "mueble_cafetera": "refrigeracion",
    "lavavajillas": "lavado", "lavautensilios": "lavado",
    "horno_combinado": "horno", "horno_conveccion": "horno",
}

app = FastAPI(
    title="Repagas - Generador de Cocinas Industriales",
    description="Webhook para recibir formulario de cliente y generar propuesta de cocina industrial.",
    version="1.0.0",
)

# allow_origins=["*"] es aceptable porque el servidor solo es accesible localmente
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

LOGIN_USER = os.getenv("REPAGAS_LOGIN_USER", "repagas")
LOGIN_PASS = os.getenv("REPAGAS_LOGIN_PASS", "concept2025")

UPLOAD_DIR = os.path.join(tempfile.gettempdir(), "repagas_uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

# Diagnostico: verificar conversores DWG disponibles al arrancar
_oda = _find_oda()
_ldwg = _find_libredwg()
print(f"[DWG] ODA File Converter: {_oda or 'NO encontrado'}")
print(f"[DWG] LibreDWG (dwg2dxf): {_ldwg or 'NO encontrado'}")
if not _oda and not _ldwg:
    print("[DWG] ADVERTENCIA: No hay conversor DWG disponible. Archivos .dwg no se podran procesar.")

# Contexto en memoria de la ultima generacion; permite que /feedback modifique la propuesta activa
_ultimo_contexto: dict = {}


def _humanizar_error(e: Exception, contexto: str = "") -> tuple[int, str]:
    """Mapea una excepcion a (status_code, mensaje_user_friendly).

    Imprime ademas la traza completa en stdout para que la veamos en logs de
    Railway sin necesidad de exponerla al usuario.
    """
    import traceback
    traceback.print_exc()  # stack completo en logs Railway

    ctx = f" durante {contexto}" if contexto else ""
    tipo = type(e).__name__
    msg = str(e)

    # Bugs internos: no exponer el detalle tecnico, dar pista para reportar
    if isinstance(e, (NameError, AttributeError, TypeError, ImportError, IndentationError, SyntaxError)):
        return (500, f"Error interno del servidor{ctx} ({tipo}). Reportalo al equipo tecnico con la fecha y la propuesta que estabas generando. Detalle: {msg[:200]}")

    # Datos del formulario incompletos / mal formados
    if isinstance(e, (KeyError, IndexError)):
        return (400, f"Falta un dato esperado en el formulario{ctx}: {msg[:200]}. Revisa que todos los campos obligatorios esten rellenados.")

    if isinstance(e, ValueError):
        return (400, f"Dato invalido en el formulario{ctx}: {msg[:200]}")

    # Archivos / planos
    if isinstance(e, FileNotFoundError):
        return (500, f"No se encontro un archivo necesario{ctx}: {msg[:200]}. Si subiste un plano, verifica que sea un .dwg o .dxf valido y no este corrupto.")

    if isinstance(e, PermissionError):
        return (500, f"Sin permisos para acceder a un archivo{ctx}. Reintenta; si persiste, contacta soporte.")

    # Red / servicios externos
    err_low = msg.lower()
    if "timeout" in err_low or "timed out" in err_low or isinstance(e, TimeoutError):
        return (504, f"El servidor tardo demasiado{ctx}. Reintenta en unos segundos.")
    if "connection" in err_low and ("refused" in err_low or "reset" in err_low):
        return (502, f"Error de conexion con un servicio externo{ctx}. Reintenta en unos segundos.")
    if "openrouter" in err_low and ("402" in msg or "payment" in err_low or "credit" in err_low):
        return (402, "OpenRouter sin creditos. Recarga saldo en https://openrouter.ai/credits o elige otro modelo desde 'Configuracion avanzada'.")
    if "openrouter" in err_low and ("429" in msg or "rate" in err_low):
        return (429, "OpenRouter rate limit alcanzado. Reintenta en unos minutos o elige otro modelo desde 'Configuracion avanzada'.")
    if "openrouter" in err_low and ("401" in msg or "unauthorized" in err_low):
        return (500, "OpenRouter rechazo la API key. Avisa al equipo tecnico para revisar la configuracion en Railway.")

    # Base de datos
    if "supabase" in err_low or "postgres" in err_low or "psycopg" in err_low or "database" in err_low:
        return (503, f"Error de conexion con la base de datos{ctx}. Reintenta en unos segundos; si persiste, contacta soporte.")

    # DXF / ezdxf
    if "dxf" in err_low or "dwg" in err_low or "ezdxf" in err_low:
        return (500, f"Error procesando el plano DXF/DWG{ctx}: {msg[:200]}. Verifica que el plano sea valido o reintenta sin plano.")

    # Generico
    return (500, f"Error inesperado{ctx} ({tipo}): {msg[:200]}")


# ─── Helpers ─────────────────────────────────────────────

def _guardar_upload(archivo: UploadFile) -> str:
    """Guarda un archivo subido en directorio temporal y devuelve la ruta."""
    ext = Path(archivo.filename).suffix.lower()
    if ext not in (".dwg", ".dxf"):
        raise HTTPException(400, f"Formato no soportado: {ext}. Usa .dwg o .dxf")

    dest = os.path.join(UPLOAD_DIR, archivo.filename)
    with open(dest, "wb") as f:
        shutil.copyfileobj(archivo.file, f)
    return dest


def _convertir_si_dwg(ruta: str) -> str:
    """
    Si el archivo es .dwg, lo convierte a .dxf. Si ya es .dxf, lo retorna.
    Lanza HTTPException(500) si la conversion falla para que el frontend se entere.
    """
    if ruta.lower().endswith(".dxf"):
        print(f"[WEBHOOK] Archivo ya es DXF, no requiere conversion")
        return ruta
    print(f"[WEBHOOK] Archivo DWG detectado, convirtiendo a DXF...")
    print(f"[WEBHOOK]   Conversor ODA disponible: {bool(_oda)}")
    print(f"[WEBHOOK]   Conversor LibreDWG disponible: {bool(_ldwg)}")
    if not _oda and not _ldwg:
        raise HTTPException(
            500,
            "No hay conversor DWG instalado en el servidor. Contacta al administrador o sube el plano como .dxf.",
        )
    try:
        dxf_path = dwg_a_dxf(ruta, output_dir=UPLOAD_DIR)
        size_kb = os.path.getsize(dxf_path) / 1024
        print(f"[WEBHOOK] DWG convertido a DXF: {dxf_path} ({size_kb:.0f}KB)")
        if size_kb < 1:
            raise HTTPException(500, "El DXF generado esta vacio. El DWG puede estar corrupto.")
        return dxf_path
    except HTTPException:
        raise
    except Exception as e:
        print(f"[WEBHOOK] ERROR convirtiendo DWG: {e}")
        raise HTTPException(500, f"No se pudo convertir el DWG: {str(e)[:200]}")


def _ejecutar_pipeline(
    formulario: FormularioCliente,
    plano_dxf: str | None = None,
    incluir_dxf: bool = False,
    descuento_general: float = 0.0,
    descuentos_individuales: dict | None = None,
    sin_descuento: list | None = None,
    openrouter_model: str | None = None,
) -> StreamingResponse:
    """Ejecuta el pipeline completo. Devuelve un ZIP con propuesta.dwg (+ dxf opcional) + PDFs + resultado.json."""

    tmp_dir = os.path.join(UPLOAD_DIR, uuid.uuid4().hex)
    os.makedirs(tmp_dir, exist_ok=True)

    try:
        print("\n[WEBHOOK] Generando propuesta con LLM (incluye RAG)...")
        propuesta = generar_propuesta_llm(formulario, openrouter_model=openrouter_model)

        # Serie se detecta desde preferencias del cliente; comensales como fallback
        prefs = (formulario.necesidades_equipamiento.preferencias_colocacion or "").lower()
        if "serie 900" in prefs or "s900" in prefs or "fondo 900" in prefs:
            serie = "900"
        elif "serie 750" in prefs or "s750" in prefs or "fondo 750" in prefs:
            serie = "750"
        else:
            serie = "900" if formulario.proyecto.comensales > 100 else "750"
        print(f"[WEBHOOK] Resolviendo equipos contra base de datos (Serie {serie})...")
        equipos_resueltos = resolver_equipos(propuesta, serie_pref=serie)

        layout_tipo = getattr(propuesta, "layout", "L")
        dxf_path = os.path.join(tmp_dir, "propuesta.dxf")
        plano_usado = False
        if plano_dxf:
            print(f"[WEBHOOK] Generando plano integrado con plano del cliente (layout={layout_tipo})...")
            dxf_path, plano_usado = generar_plano_integrado(equipos_resueltos, plano_dxf, filepath=dxf_path, layout_tipo=layout_tipo)
            print(f"[WEBHOOK] Plano cliente usado: {plano_usado}")
        else:
            print(f"[WEBHOOK] Generando plano DXF standalone (layout={layout_tipo})...")
            dxf_path = generar_plano(equipos_resueltos, filepath=dxf_path, layout_tipo=layout_tipo)
        png_path = dxf_path.replace(".dxf", ".png")

        total_pvp = sum((e.pvp_eur or 0) * e.cantidad for e in equipos_resueltos)
        resultado = {
            "proyecto": propuesta.nombre_proyecto,
            "layout": propuesta.layout,
            "equipos": [
                {
                    "modelo": e.modelo,
                    "tipo": e.tipo,
                    "ancho_mm": e.ancho_mm,
                    "fondo_mm": e.fondo_mm,
                    "pvp_eur": e.pvp_eur,
                    "cantidad": e.cantidad,
                    "zona": e.zona,
                    "serie": e.serie,
                }
                for e in equipos_resueltos
            ],
            "total_equipos": len(equipos_resueltos),
            "total_pvp_eur": round(total_pvp, 2),
            "notas_llm": propuesta.notas,
            "plano_cliente_usado": plano_usado,
            "tenia_plano_cliente": bool(plano_dxf),
        }

        imprimir_resumen(formulario, equipos_resueltos, dxf_path)

        try:
            conn = get_db_connection()
            cur = conn.cursor()
            cur.execute(
                """INSERT INTO historico_propuestas
                   (nombre_proyecto, tipo_negocio, comensales, total_equipos,
                    total_pvp_eur, layout, notas_llm, formulario, propuesta, equipos)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s::jsonb, %s::jsonb, %s::jsonb)""",
                (
                    propuesta.nombre_proyecto or "Sin nombre",
                    formulario.proyecto.tipo_negocio,
                    formulario.proyecto.comensales,
                    len(equipos_resueltos),
                    round(total_pvp, 2),
                    propuesta.layout,
                    propuesta.notas or "",
                    json.dumps(formulario.model_dump(), ensure_ascii=False),
                    propuesta.model_dump_json(),
                    json.dumps(resultado["equipos"], ensure_ascii=False),
                ),
            )
            conn.commit()
            cur.close(); conn.close()
        except Exception as e:
            print(f"[WEBHOOK] WARN: No se pudo guardar en historico: {e}")

        _ultimo_contexto.clear()
        _ultimo_contexto["formulario"] = formulario
        _ultimo_contexto["propuesta"] = propuesta
        _ultimo_contexto["equipos_resueltos"] = equipos_resueltos
        _ultimo_contexto["plano_dxf"] = plano_dxf
        _ultimo_contexto["descuento_general"] = descuento_general
        _ultimo_contexto["descuentos_individuales"] = descuentos_individuales or {}
        _ultimo_contexto["sin_descuento"] = sin_descuento or []

        nombre_proy = propuesta.nombre_proyecto or ""
        formulario_dict = formulario.model_dump()

        from features.planos.posicionar import EquipoPosicionado
        equipos_pos = [
            EquipoPosicionado(
                modelo=e.modelo, tipo=e.tipo, zona=e.zona,
                ancho_mm=e.ancho_mm, fondo_mm=e.fondo_mm, alto_mm=e.alto_mm,
                pvp_eur=e.pvp_eur, serie=e.serie, cantidad=1,
                x=0, y=0, rotation=0, corners=None, wall_side="north",
            )
            for e in equipos_resueltos
            for _ in range(e.cantidad)
        ]

        pdf_prospeccion = os.path.join(tmp_dir, "prospeccion.pdf")
        try:
            generar_pdf_prospeccion(formulario_dict, nombre_proy, pdf_prospeccion)
        except Exception as e:
            print(f"[WEBHOOK] WARN: PDF prospeccion fallo: {e}")
            pdf_prospeccion = None

        pdf_presupuesto = os.path.join(tmp_dir, "presupuesto.pdf")
        try:
            generar_pdf_presupuesto(
                equipos_pos, nombre_proy, formulario_dict, pdf_presupuesto,
                descuento_general=descuento_general,
                descuentos_individuales=descuentos_individuales or {},
                sin_descuento=sin_descuento or [],
            )
        except Exception as e:
            print(f"[WEBHOOK] WARN: PDF presupuesto fallo: {e}")
            pdf_presupuesto = None

        # Tabla de equipos en DXF/DWG (para pegar en el plano de AutoCAD)
        tabla_dxf_path = os.path.join(tmp_dir, "tabla_equipos.dxf")
        tabla_dwg_path = None
        try:
            from features.documentos.tabla import generar_tabla_equipos_dxf
            generar_tabla_equipos_dxf(equipos_resueltos, nombre_proy, tabla_dxf_path)
        except Exception as e:
            print(f"[WEBHOOK] WARN: No se pudo generar tabla DXF: {e}")
            tabla_dxf_path = None

        # Convertir DXF -> DWG (formato nativo AutoCAD)
        dwg_path = None
        try:
            from features.planos.conversion import dxf_a_dwg
            dwg_path = dxf_a_dwg(dxf_path, output_dir=tmp_dir)
        except Exception as e:
            print(f"[WEBHOOK] WARN: No se pudo generar DWG: {e}")

        if tabla_dxf_path and os.path.isfile(tabla_dxf_path):
            try:
                from features.planos.conversion import dxf_a_dwg
                tabla_dwg_path = dxf_a_dwg(tabla_dxf_path, output_dir=tmp_dir)
            except Exception as e:
                print(f"[WEBHOOK] WARN: No se pudo convertir tabla a DWG: {e}")

        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("resultado.json", json.dumps(resultado, ensure_ascii=False, indent=2))
            if dwg_path and os.path.isfile(dwg_path):
                zf.write(dwg_path, "propuesta.dwg")
            if incluir_dxf and os.path.isfile(dxf_path):
                zf.write(dxf_path, "propuesta.dxf")
            elif not dwg_path and os.path.isfile(dxf_path):
                # Si no hay DWG disponible, incluir DXF como fallback
                zf.write(dxf_path, "propuesta.dxf")
            if os.path.isfile(png_path):
                zf.write(png_path, "propuesta.png")
            if pdf_prospeccion and os.path.isfile(pdf_prospeccion):
                zf.write(pdf_prospeccion, "prospeccion.pdf")
            if pdf_presupuesto and os.path.isfile(pdf_presupuesto):
                zf.write(pdf_presupuesto, "presupuesto.pdf")
            if tabla_dwg_path and os.path.isfile(tabla_dwg_path):
                zf.write(tabla_dwg_path, "tabla_equipos.dwg")
            if tabla_dxf_path and os.path.isfile(tabla_dxf_path):
                # incluir DXF tambien (fallback si DWG fallo, o por si acaso)
                zf.write(tabla_dxf_path, "tabla_equipos.dxf")

        zip_buffer.seek(0)

        return StreamingResponse(
            zip_buffer,
            media_type="application/zip",
            headers={"Content-Disposition": "attachment; filename=propuesta.zip"},
        )

    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


# ─── Endpoints ───────────────────────────────────────────

# Mapa de palabras-zona para el parser local del Excel (espejo del que vive en
# formulario-repagas/app.js).
_EXCEL_ZONA_MAP = [
    ("COCCION",       "coccion"),
    ("LAVADO",        "lavado"),
    ("ALMACEN",       "refrigeracion"),
    ("FRIO",          "refrigeracion"),
    ("REFRIGERACION", "refrigeracion"),
    ("HORNO",         "horno"),
    ("BARRA",         None),   # se ignora
]


def _normalizar_para_zona(s: str) -> str:
    """NFD + strip de diacriticos + upper + trim. Espejo de la helper del frontend."""
    import unicodedata
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.upper().strip()


def _parsear_excel_local(wb) -> dict:
    """Parser local del Excel sin IA. Detecta cabeceras 'ZONA X' con word boundary
    (evita 'PRELAVADO' matchee 'LAVADO') y extrae equipos por zona. Mismas reglas
    que el frontend en formulario-repagas/app.js (importEquiposFromExcelRows)."""
    import re

    agrupado = {"coccion": [], "refrigeracion": [], "lavado": [], "horno": []}
    current_zone = None

    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            text_cells = []
            num_cells = []
            for cell in row:
                if cell is None:
                    continue
                if isinstance(cell, (int, float)):
                    num_cells.append(cell)
                else:
                    t = str(cell).strip()
                    if t:
                        text_cells.append(t)
            if not text_cells:
                continue

            # 1) Detectar cabecera de zona usando word boundary
            is_header = False
            for text in text_cells:
                tn = _normalizar_para_zona(text)
                for key, zone in _EXCEL_ZONA_MAP:
                    if re.search(rf"\b{re.escape(key)}\b", tn):
                        current_zone = zone
                        is_header = True
                        break
                if is_header:
                    break
            if is_header:
                continue
            if current_zone is None:
                continue

            # 2) Nombre del equipo: texto mas largo (descarta "UNIDADES" / "UDS")
            candidatos = [
                t for t in text_cells
                if t.upper() not in ("UNIDADES", "CANTIDAD", "UDS") and len(t) > 2
            ]
            if not candidatos:
                continue
            raw_name = max(candidatos, key=len)
            m = re.match(r"^\d+(?:\.\d+)?\s+([\s\S]+)", raw_name)
            nombre = (m.group(1) if m else raw_name).strip()
            if not nombre:
                continue

            # 3) Cantidad: primer numero de la fila (>=1)
            cantidad = max(1, int(round(num_cells[0]))) if num_cells else 1
            agrupado[current_zone].append({"nombre": nombre, "cantidad": cantidad})

    return agrupado


@app.post("/parsear-excel")
async def parsear_excel(archivo: UploadFile = File(...)):
    """
    Lee un Excel de equipos y devuelve los equipos agrupados por zona.
    Usa parseo heuristico (cabeceras ZONA X) y pide a la IA desambiguar si hace falta.
    """
    ext = os.path.splitext(archivo.filename)[1].lower()
    if ext not in (".xlsx", ".xls"):
        raise HTTPException(400, f"Formato no soportado: {ext}. Usa .xlsx")

    tmp_path = os.path.join(UPLOAD_DIR, f"excel_{uuid.uuid4().hex}{ext}")
    try:
        with open(tmp_path, "wb") as f:
            f.write(await archivo.read())

        import openpyxl
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
        filas_texto = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                celdas = [str(c).strip() for c in row if c is not None and str(c).strip()]
                if celdas:
                    filas_texto.append(" | ".join(celdas))
        texto_completo = "\n".join(filas_texto)

        if not texto_completo.strip():
            raise HTTPException(400, "El Excel esta vacio")

        from features.propuestas.llm import invocar_llm_con_rotacion
        from langchain_core.messages import SystemMessage, HumanMessage
        from pydantic import BaseModel, Field

        class EquipoExcel(BaseModel):
            nombre: str
            cantidad: int = 1
            zona: str = Field(description="Una de: coccion, refrigeracion, lavado, horno")

        class ResultadoExcel(BaseModel):
            equipos: list[EquipoExcel]

        messages = [
            SystemMessage(content=(
                "Eres un asistente que extrae listas de equipamiento de cocina industrial de archivos Excel. "
                "Recibes el contenido texto del Excel y devuelves una lista estructurada de equipos con su zona. "
                "Zonas validas: 'coccion' (cocinas, hornos, freidoras, paelleros, barbacoas, planchas, fry-tops, "
                "campanas, mesas de trabajo de coccion), 'refrigeracion' (armarios frigorificos, mesas refrigeradas, "
                "armarios congelacion, abatidores, botelleros), 'lavado' (lavavajillas, mesas de salida/prelavado, "
                "fregaderos, GS, estanterias de lavado), 'horno' (si esta separado, si no usar coccion). "
                "IMPORTANTE: Ignora cabeceras/secciones (textos como 'ZONA COCCION', 'UNIDADES', filas vacias). "
                "Extrae solo equipos reales con su nombre limpio (sin codigos como '1.01' o '2.03'). "
                "La cantidad suele estar en una columna aparte; si ves 'x3' o el numero 3 en una fila, son 3 unidades."
            )),
            HumanMessage(content=(
                f"Contenido del Excel:\n\n{texto_completo}\n\n"
                f"Devuelve la lista estructurada de equipos."
            )),
        ]

        print(f"[EXCEL] Analizando con IA: {archivo.filename} ({len(texto_completo)} chars)")
        try:
            resultado = invocar_llm_con_rotacion(messages, structured_cls=ResultadoExcel)
        except Exception as e:
            print(f"[EXCEL] Excepcion en IA: {type(e).__name__}: {str(e)[:200]}")
            resultado = None

        if resultado:
            agrupado = {"coccion": [], "refrigeracion": [], "lavado": [], "horno": []}
            for eq in resultado.equipos:
                zona = eq.zona.lower()
                if zona not in agrupado:
                    zona = "coccion"
                agrupado[zona].append({"nombre": eq.nombre, "cantidad": max(1, eq.cantidad)})

            total = sum(len(v) for v in agrupado.values())
            print(f"[EXCEL] IA extrajo {total} equipos: " +
                  ", ".join(f"{k}={len(v)}" for k, v in agrupado.items() if v))
            return agrupado

        # Fallback local sin IA: parser openpyxl + cabeceras de zona con word
        # boundary (mismas reglas que el frontend). Asi /parsear-excel sigue
        # devolviendo resultado util cuando OpenRouter no responde.
        print("[EXCEL] IA no respondio; usando parser local con word boundary...")
        agrupado = _parsear_excel_local(wb)
        total = sum(len(v) for v in agrupado.values())
        print(f"[EXCEL] Parser local extrajo {total} equipos: " +
              ", ".join(f"{k}={len(v)}" for k, v in agrupado.items() if v))
        return agrupado
    except HTTPException:
        raise
    except Exception as e:
        code, msg = _humanizar_error(e, contexto="parseo del Excel")
        raise HTTPException(code, msg)
    finally:
        if os.path.isfile(tmp_path):
            os.remove(tmp_path)


@app.get("/diagnostico")
async def diagnostico():
    """Reporta el estado de conversores DWG y otras dependencias criticas."""
    import glob as _glob
    from features.planos.conversion import _find_libredwg as _find
    dxf2dwg = _find("dxf2dwg")
    libredwg_bins = []
    for d in ["/usr/bin", "/usr/local/bin"]:
        libredwg_bins.extend(_glob.glob(f"{d}/*dwg*") + _glob.glob(f"{d}/*dxf*"))
    return {
        "oda_converter": _oda or None,
        "libredwg_dwg2dxf": _ldwg or None,
        "libredwg_dxf2dwg": dxf2dwg or None,
        "puede_procesar_dwg_entrada": bool(_oda or _ldwg),
        "puede_generar_dwg_salida": bool(_oda or dxf2dwg),
        "binarios_disponibles": sorted(set(libredwg_bins)),
    }


# Cache simple en memoria para /openrouter/modelos (TTL 1 hora)
_OPENROUTER_MODELOS_CACHE: dict = {"ts": 0, "data": None}


@app.get("/openrouter/modelos")
async def listar_modelos_openrouter():
    """Lista los modelos disponibles en OpenRouter para que el frontend pueda
    mostrarlos en el desplegable de 'Configuracion avanzada'. Cachea la respuesta
    1 hora para no machacar la API de OpenRouter."""
    import time
    import urllib.request
    import urllib.error
    from features.propuestas.llm import OPENROUTER_MODEL_DEFAULT

    ahora = time.time()
    if _OPENROUTER_MODELOS_CACHE["data"] and (ahora - _OPENROUTER_MODELOS_CACHE["ts"]) < 3600:
        return _OPENROUTER_MODELOS_CACHE["data"]

    req = urllib.request.Request("https://openrouter.ai/api/v1/models")
    try:
        with urllib.request.urlopen(req, timeout=15) as r:
            data = json.loads(r.read().decode())
    except Exception as e:
        # Fallback: lista hardcodeada de modelos populares
        return {
            "default": OPENROUTER_MODEL_DEFAULT,
            "modelos": _MODELOS_FALLBACK,
            "fuente": "fallback_local",
            "error": str(e)[:200],
        }

    modelos_raw = data.get("data", []) if isinstance(data, dict) else []
    modelos = []
    for m in modelos_raw:
        mid = m.get("id")
        if not mid:
            continue
        pricing = m.get("pricing", {}) or {}
        prompt_p = pricing.get("prompt", "0")
        comp_p = pricing.get("completion", "0")
        modelos.append({
            "id": mid,
            "name": m.get("name") or mid,
            "context_length": m.get("context_length"),
            "prompt_price": prompt_p,
            "completion_price": comp_p,
            "vendor": mid.split("/")[0] if "/" in mid else "",
        })

    # Orden: default primero, luego modelos :free agrupados por vendor (google/
    # anthropic/openai/otros), luego modelos de pago igual ordenados. Asi quien no
    # tiene saldo en OpenRouter ve los gratis priorizados despues del default.
    def _orden(m):
        vendor_priority = {"google": 0, "anthropic": 1, "openai": 2, "deepseek": 3, "meta-llama": 4}
        is_default = 0 if m["id"] == OPENROUTER_MODEL_DEFAULT else 1
        is_paid = 0 if m["id"].endswith(":free") else 1
        v = vendor_priority.get(m["vendor"], 9)
        return (is_default, is_paid, v, m["id"])

    modelos.sort(key=_orden)

    resp = {
        "default": OPENROUTER_MODEL_DEFAULT,
        "modelos": modelos,
        "fuente": "openrouter_api",
    }
    _OPENROUTER_MODELOS_CACHE["ts"] = ahora
    _OPENROUTER_MODELOS_CACHE["data"] = resp
    return resp


# Lista pequena de fallback si OpenRouter no responde. Free primero (porque el
# default es free), luego algunos de pago populares para clientes con saldo.
_MODELOS_FALLBACK = [
    {"id": "deepseek/deepseek-v4-flash:free",        "name": "DeepSeek V4 Flash (gratis)",   "vendor": "deepseek"},
    {"id": "openai/gpt-oss-120b:free",               "name": "OpenAI gpt-oss 120B (gratis)", "vendor": "openai"},
    {"id": "z-ai/glm-4.5-air:free",                  "name": "GLM 4.5 Air (gratis)",         "vendor": "z-ai"},
    {"id": "meta-llama/llama-3.3-70b-instruct:free", "name": "Llama 3.3 70B (gratis)",       "vendor": "meta-llama"},
    {"id": "google/gemini-3-flash-preview",          "name": "Gemini 3 Flash (Preview)",     "vendor": "google"},
    {"id": "google/gemini-2.5-flash",                "name": "Gemini 2.5 Flash",             "vendor": "google"},
    {"id": "google/gemini-2.5-pro",                  "name": "Gemini 2.5 Pro",               "vendor": "google"},
    {"id": "anthropic/claude-sonnet-4.5",            "name": "Claude Sonnet 4.5",            "vendor": "anthropic"},
    {"id": "openai/gpt-5",                           "name": "GPT-5",                        "vendor": "openai"},
]


@app.get("/diagnostico/llm/saldo")
async def diagnostico_llm_saldo():
    """Consulta el saldo / limite / uso de la OPENROUTER_API_KEY contra los
    endpoints publicos de OpenRouter. NO devuelve la API key (solo metadata).

    Util para verificar si OpenRouter sigue teniendo creditos sin necesidad
    de leer la variable de entorno en Railway.
    """
    import urllib.request
    import urllib.error
    from features.propuestas.llm import OPENROUTER_KEY, OPENROUTER_BASE_URL

    if not OPENROUTER_KEY:
        return {"ok": False, "error": "OPENROUTER_API_KEY no configurada en el entorno."}

    headers = {"Authorization": f"Bearer {OPENROUTER_KEY}"}

    def _fetch(path):
        req = urllib.request.Request(f"{OPENROUTER_BASE_URL}{path}", headers=headers)
        try:
            with urllib.request.urlopen(req, timeout=15) as r:
                import json as _json
                return r.status, _json.loads(r.read().decode())
        except urllib.error.HTTPError as e:
            try:
                body = _json.loads(e.read().decode())
            except Exception:
                body = {"raw": str(e)}
            return e.code, body
        except Exception as e:
            return 0, {"error": str(e)[:200]}

    # /credits devuelve {data: {total_credits, total_usage}}
    code_credits, data_credits = _fetch("/credits")
    # /auth/key devuelve {data: {label, usage, limit, is_free_tier, rate_limit, ...}}
    code_key, data_key = _fetch("/auth/key")

    def _redact(d):
        # Defensivo: si por algun motivo OpenRouter devolviera la key, no la propagamos
        if isinstance(d, dict):
            return {k: ("***" if "key" in k.lower() and isinstance(v, str) and len(v) > 12 else _redact(v)) for k, v in d.items()}
        if isinstance(d, list):
            return [_redact(x) for x in d]
        return d

    saldo = {}
    if code_credits == 200 and isinstance(data_credits, dict):
        d = data_credits.get("data", data_credits)
        total = d.get("total_credits")
        usage = d.get("total_usage")
        if total is not None and usage is not None:
            saldo["credits_total"] = total
            saldo["credits_used"] = usage
            saldo["credits_restantes"] = round(float(total) - float(usage), 6)

    if code_key == 200 and isinstance(data_key, dict):
        d = data_key.get("data", data_key)
        saldo["label"] = d.get("label")
        saldo["is_free_tier"] = d.get("is_free_tier")
        saldo["usage_key"] = d.get("usage")
        saldo["limit_key"] = d.get("limit")
        saldo["rate_limit"] = d.get("rate_limit")

    return {
        "ok": code_credits == 200 and code_key == 200,
        "openrouter_credits_status": code_credits,
        "openrouter_auth_key_status": code_key,
        **saldo,
        # En caso de error de OpenRouter, propagamos los detalles redactados
        "raw_credits": _redact(data_credits) if code_credits != 200 else None,
        "raw_auth_key": _redact(data_key) if code_key != 200 else None,
    }


@app.get("/diagnostico/llm")
async def diagnostico_llm():
    """Verifica que el LLM esta configurado y responde. Util para confirmar que
    OpenRouter funciona (en produccion es el unico proveedor) y leer el ultimo
    motivo de fallo si el usuario reporta que vio el fallback."""
    from features.propuestas.llm import (
        OPENROUTER_KEY, OPENROUTER_MODEL, OPENROUTER_BASE_URL, GEMINI_KEYS,
        invocar_llm_con_rotacion, obtener_ultimo_fallo_llm,
    )
    config = {
        "openrouter_configurado": bool(OPENROUTER_KEY),
        "openrouter_model": OPENROUTER_MODEL,
        "openrouter_base_url": OPENROUTER_BASE_URL,
        "gemini_keys_count": len(GEMINI_KEYS),
        "ultimo_fallo": obtener_ultimo_fallo_llm(),
    }
    # Test ligero: pedir un "ok" muy corto al LLM
    try:
        ping = invocar_llm_con_rotacion(
            [
                {"role": "system", "content": "Responde unicamente con la palabra 'ok'."},
                {"role": "human", "content": "Ping"},
            ],
            structured_cls=None,
            max_reintentos=0,
        )
        if ping is None:
            return {
                **config,
                "llm_responde": False,
                "ultimo_fallo": obtener_ultimo_fallo_llm(),
            }
        contenido = getattr(ping, "content", str(ping))
        return {
            **config,
            "llm_responde": True,
            "respuesta": str(contenido)[:200],
        }
    except Exception as e:
        return {
            **config,
            "llm_responde": False,
            "error_inesperado": str(e)[:300],
        }


@app.post("/login")
async def login(data: dict):
    """Valida credenciales simples para acceso al formulario."""
    user = data.get("user", "")
    password = data.get("pass", "")
    if user == LOGIN_USER and password == LOGIN_PASS:
        return {"ok": True}
    raise HTTPException(status_code=401, detail="Credenciales incorrectas")


@app.get("/historico")
async def listar_historico(limit: int = 50):
    """Lista las propuestas generadas (mas recientes primero)."""
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            """SELECT id, nombre_proyecto, tipo_negocio, comensales,
                      total_equipos, total_pvp_eur, layout, creado_en
               FROM historico_propuestas
               ORDER BY creado_en DESC
               LIMIT %s""",
            (limit,),
        )
        rows = cur.fetchall()
        cur.close(); conn.close()
        return [
            {
                "id": str(r[0]),
                "nombre_proyecto": r[1],
                "tipo_negocio": r[2],
                "comensales": r[3],
                "total_equipos": r[4],
                "total_pvp_eur": float(r[5]) if r[5] else 0,
                "layout": r[6],
                "creado_en": r[7].isoformat() if r[7] else None,
            }
            for r in rows
        ]
    except Exception as e:
        raise HTTPException(500, f"Error consultando historico: {str(e)[:200]}")


@app.get("/historico/{prop_id}")
async def obtener_propuesta(prop_id: str):
    """Devuelve una propuesta del historico con todos sus datos."""
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            """SELECT id, nombre_proyecto, tipo_negocio, comensales,
                      total_equipos, total_pvp_eur, layout, notas_llm,
                      formulario, propuesta, equipos, creado_en
               FROM historico_propuestas
               WHERE id = %s""",
            (prop_id,),
        )
        r = cur.fetchone()
        cur.close(); conn.close()
        if not r:
            raise HTTPException(404, "Propuesta no encontrada")
        return {
            "id": str(r[0]),
            "nombre_proyecto": r[1],
            "tipo_negocio": r[2],
            "comensales": r[3],
            "total_equipos": r[4],
            "total_pvp_eur": float(r[5]) if r[5] else 0,
            "layout": r[6],
            "notas_llm": r[7],
            "formulario": r[8],
            "propuesta": r[9],
            "equipos": r[10],
            "creado_en": r[11].isoformat() if r[11] else None,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Error: {str(e)[:200]}")


@app.delete("/historico/{prop_id}")
async def eliminar_propuesta(prop_id: str):
    """Elimina una propuesta del historico."""
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM historico_propuestas WHERE id = %s", (prop_id,))
        deleted = cur.rowcount
        conn.commit()
        cur.close(); conn.close()
        if deleted == 0:
            raise HTTPException(404, "Propuesta no encontrada")
        return {"ok": True, "eliminados": deleted}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Error: {str(e)[:200]}")


@app.get("/admin/rag")
async def listar_documentos_rag(limit: int = 100):
    """Lista los documentos indexados en el RAG."""
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            """SELECT id, titulo, tipo_archivo, categoria, num_chunks, procesado
               FROM documentos_rag
               ORDER BY titulo
               LIMIT %s""",
            (limit,),
        )
        rows = cur.fetchall()
        cur.close(); conn.close()
        return [
            {
                "id": str(r[0]),
                "titulo": r[1],
                "tipo_archivo": r[2],
                "categoria": r[3],
                "num_chunks": r[4],
                "procesado": r[5],
            }
            for r in rows
        ]
    except Exception as e:
        raise HTTPException(500, f"Error: {str(e)[:200]}")


@app.post("/admin/rag/upload")
async def subir_documento_rag(
    archivo: UploadFile = File(...),
    categoria: str = Form("general"),
):
    """
    Sube un documento (PDF, DOCX, PPTX) y lo indexa en el RAG.
    Extrae texto, genera chunks, calcula embeddings y los guarda en Supabase.
    """
    ext = os.path.splitext(archivo.filename)[1].lower()
    if ext not in (".pdf", ".pptx", ".pptm", ".docx"):
        raise HTTPException(400, f"Formato no soportado: {ext}. Usa PDF, PPTX o DOCX.")

    tmp_path = os.path.join(UPLOAD_DIR, f"rag_{uuid.uuid4().hex}{ext}")
    try:
        with open(tmp_path, "wb") as f:
            f.write(await archivo.read())

        from features.rag.pipeline import (
            extraer_texto_pdf, extraer_texto_pptx, extraer_texto_docx,
            chunkear_texto, generar_embeddings, almacenar_documento,
            detectar_categoria,
        )

        if ext == ".pdf":
            texto = extraer_texto_pdf(tmp_path)
        elif ext in (".pptx", ".pptm"):
            texto = extraer_texto_pptx(tmp_path)
        else:
            texto = extraer_texto_docx(tmp_path)

        if not texto or len(texto.strip()) < 50:
            raise HTTPException(400, "No se pudo extraer texto del documento (o texto muy corto)")

        chunks = chunkear_texto(texto)
        if not chunks:
            raise HTTPException(400, "No se generaron chunks del documento")

        print(f"[RAG] Procesando '{archivo.filename}': {len(texto)} chars, {len(chunks)} chunks")
        embeddings = generar_embeddings(chunks)

        cat = categoria if categoria != "auto" else detectar_categoria(archivo.filename)
        titulo = os.path.splitext(archivo.filename)[0]
        tipo_archivo = ext.replace(".", "")

        conn = get_db_connection()
        conn.autocommit = False
        cur = conn.cursor()
        doc_id = almacenar_documento(
            cur, titulo, tipo_archivo, f"upload://{archivo.filename}",
            cat, chunks, embeddings,
        )
        conn.commit()
        cur.close(); conn.close()

        return {
            "ok": True,
            "id": str(doc_id),
            "titulo": titulo,
            "categoria": cat,
            "chunks": len(chunks),
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"[RAG] Error: {e}")
        raise HTTPException(500, f"Error procesando documento: {str(e)[:200]}")
    finally:
        if os.path.isfile(tmp_path):
            os.remove(tmp_path)


@app.delete("/admin/rag/{doc_id}")
async def eliminar_documento_rag(doc_id: str):
    """Elimina un documento del RAG (y sus chunks)."""
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM chunks_rag WHERE documento_id = %s", (doc_id,))
        cur.execute("DELETE FROM documentos_rag WHERE id = %s", (doc_id,))
        deleted = cur.rowcount
        conn.commit()
        cur.close(); conn.close()
        if deleted == 0:
            raise HTTPException(404, "Documento no encontrado")
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Error: {str(e)[:200]}")


@app.get("/admin/textos")
async def listar_textos_config():
    """Devuelve todos los textos configurables de los PDFs (clave -> valor)."""
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT clave, valor FROM textos_config")
        rows = cur.fetchall()
        cur.close(); conn.close()
        return {r[0]: r[1] for r in rows}
    except Exception as e:
        raise HTTPException(500, f"Error: {str(e)[:200]}")


@app.put("/admin/textos/{clave}")
async def actualizar_texto(clave: str, data: dict):
    """Actualiza un texto configurable (upsert)."""
    valor = data.get("valor", "")
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            """INSERT INTO textos_config (clave, valor, actualizado_en)
               VALUES (%s, %s, NOW())
               ON CONFLICT (clave) DO UPDATE
               SET valor = EXCLUDED.valor, actualizado_en = NOW()""",
            (clave, valor),
        )
        conn.commit()
        cur.close(); conn.close()
        return {"ok": True, "clave": clave}
    except Exception as e:
        raise HTTPException(500, f"Error: {str(e)[:200]}")


def cargar_textos_config() -> dict:
    """Helper para leer textos_config desde los generadores de PDF. Tolera fallos."""
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT clave, valor FROM textos_config")
        rows = cur.fetchall()
        cur.close(); conn.close()
        return {r[0]: r[1] for r in rows}
    except Exception:
        return {}


@app.get("/admin/libreria")
async def listar_bloques_libreria():
    """Lista los bloques CAD actualmente disponibles en la libreria."""
    try:
        import json as _json
        _BASE = os.path.dirname(os.path.abspath(__file__))
        mapa_path = os.path.join(_BASE, "data", "bloque_map.json")
        if not os.path.isfile(mapa_path):
            return {"total": 0, "bloques": []}
        with open(mapa_path, encoding="utf-8") as f:
            mapa = _json.load(f)
        bloques = [
            {
                "nombre": nombre,
                "width_mm": info.get("width_mm"),
                "depth_mm": info.get("depth_mm"),
            }
            for nombre, info in sorted(mapa.items())
        ]
        return {"total": len(bloques), "bloques": bloques}
    except Exception as e:
        raise HTTPException(500, f"Error: {str(e)[:200]}")


@app.post("/admin/libreria/upload")
async def subir_libreria_cad(archivo: UploadFile = File(...)):
    """
    Sube un DXF o DWG con nuevos bloques y los anade a la libreria maestra
    (libreria_bloques.dxf + bloque_map.json). El motor de IA podra usarlos
    en futuras propuestas.
    """
    ext = os.path.splitext(archivo.filename)[1].lower()
    if ext not in (".dxf", ".dwg"):
        raise HTTPException(400, f"Formato no soportado: {ext}. Usa .dxf o .dwg")

    tmp_path = os.path.join(UPLOAD_DIR, f"lib_{uuid.uuid4().hex}{ext}")
    try:
        with open(tmp_path, "wb") as f:
            f.write(await archivo.read())

        origen_dxf = _convertir_si_dwg(tmp_path) if ext == ".dwg" else tmp_path

        import ezdxf
        import json as _json
        _BASE = os.path.dirname(os.path.abspath(__file__))
        mapa_path = os.path.join(_BASE, "data", "bloque_map.json")
        libreria_path = os.path.join(_BASE, "data", "libreria_bloques.dxf")

        try:
            from ezdxf import recover
            origen_doc, _ = recover.readfile(origen_dxf)
        except Exception:
            origen_doc = ezdxf.readfile(origen_dxf)

        if os.path.isfile(libreria_path):
            libreria_doc = ezdxf.readfile(libreria_path)
        else:
            libreria_doc = ezdxf.new("R2018")

        with open(mapa_path, encoding="utf-8") as f:
            mapa = _json.load(f) if os.path.isfile(mapa_path) else {}

        from ezdxf.addons import Importer
        importer = Importer(origen_doc, libreria_doc)

        bloques_nuevos = []
        bloques_existentes = []
        for block in origen_doc.blocks:
            bname = block.name
            if bname.startswith("*") or bname.upper() in ("$MODEL_SPACE", "$PAPER_SPACE"):
                continue
            if bname in mapa:
                bloques_existentes.append(bname)
                continue
            try:
                extmin = [0, 0, 0]
                extmax = [0, 0, 0]
                try:
                    from ezdxf import bbox as _bbox
                    cache = _bbox.extents(block)
                    if cache.has_data:
                        extmin = [cache.extmin.x, cache.extmin.y, cache.extmin.z]
                        extmax = [cache.extmax.x, cache.extmax.y, cache.extmax.z]
                except Exception:
                    pass
                w = extmax[0] - max(0, extmin[0])
                d = extmax[1] - max(0, extmin[1])
                if w < 50 or d < 50:
                    continue  # bloque demasiado pequeno
                importer.import_block(bname)
                mapa[bname] = {
                    "width_mm": round(w, 2),
                    "depth_mm": round(d, 2),
                    "extmin": extmin,
                    "extmax": extmax,
                }
                bloques_nuevos.append(bname)
            except Exception as e:
                print(f"[LIBRERIA] No se pudo importar {bname}: {e}")

        importer.finalize()
        libreria_doc.saveas(libreria_path)
        with open(mapa_path, "w", encoding="utf-8") as f:
            _json.dump(mapa, f, indent=2, ensure_ascii=False)

        # Invalidar cache en memoria de bloque_map para que se recargue
        try:
            from features.planos.integrar import _bloque_map
            _bloque_map.clear()
            _bloque_map.update(mapa)
        except Exception:
            pass

        return {
            "ok": True,
            "bloques_nuevos": bloques_nuevos,
            "bloques_existentes": bloques_existentes,
            "total_libreria": len(mapa),
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"[LIBRERIA] Error: {e}")
        raise HTTPException(500, f"Error procesando libreria: {str(e)[:300]}")
    finally:
        if os.path.isfile(tmp_path):
            os.remove(tmp_path)


@app.get("/admin/equipos")
async def listar_equipos_admin(limit: int = 500):
    """Lista equipos de la BD para administracion."""
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            """SELECT e.id, e.modelo, e.tipo, e.alimentacion, e.ancho_mm,
                      e.fondo_mm, e.alto_mm, e.pvp_eur, s.nombre as serie, e.activo
               FROM equipos e
               LEFT JOIN series s ON e.serie_id = s.id
               WHERE e.activo = TRUE
               ORDER BY e.tipo, e.modelo
               LIMIT %s""",
            (limit,),
        )
        rows = cur.fetchall()
        cur.close(); conn.close()
        return [
            {
                "id": r[0],
                "modelo": r[1],
                "tipo": r[2],
                "alimentacion": r[3],
                "ancho_mm": r[4],
                "fondo_mm": r[5],
                "alto_mm": r[6],
                "pvp_eur": float(r[7]) if r[7] else None,
                "serie": r[8],
                "activo": r[9],
            }
            for r in rows
        ]
    except Exception as e:
        raise HTTPException(500, f"Error: {str(e)[:200]}")


@app.put("/admin/equipos/{equipo_id}")
async def actualizar_equipo(equipo_id: str, data: dict):
    """Actualiza un equipo (precio, dimensiones, etc.). equipo_id es UUID."""
    campos_validos = {"modelo", "tipo", "ancho_mm", "fondo_mm", "alto_mm", "pvp_eur", "alimentacion", "activo"}
    updates = {k: v for k, v in data.items() if k in campos_validos}
    if not updates:
        raise HTTPException(400, "Sin campos validos para actualizar")
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        sets = ", ".join(f"{k} = %s" for k in updates.keys())
        params = list(updates.values()) + [equipo_id]
        cur.execute(f"UPDATE equipos SET {sets} WHERE id = %s", params)
        conn.commit()
        updated = cur.rowcount
        cur.close(); conn.close()
        if updated == 0:
            raise HTTPException(404, "Equipo no encontrado")
        return {"ok": True, "actualizado": equipo_id}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Error: {str(e)[:200]}")


@app.post("/admin/equipos")
async def crear_equipo(data: dict):
    """Crea un nuevo equipo en el catalogo manualmente."""
    requeridos = ["modelo", "tipo"]
    for k in requeridos:
        if not data.get(k):
            raise HTTPException(400, f"Campo requerido: {k}")

    campos_permitidos = {"modelo", "tipo", "ancho_mm", "fondo_mm", "alto_mm",
                         "pvp_eur", "alimentacion", "serie_id"}
    filtered = {k: v for k, v in data.items() if k in campos_permitidos and v not in (None, "")}
    filtered.setdefault("alimentacion", "gas")
    filtered["activo"] = True
    # modelo_normalizado es NOT NULL; se genera en minusculas sin caracteres especiales
    import re as _re
    filtered["modelo_normalizado"] = _re.sub(r"[^a-z0-9]", "", filtered["modelo"].lower())

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cols = ", ".join(filtered.keys())
        placeholders = ", ".join(["%s"] * len(filtered))
        cur.execute(
            f"INSERT INTO equipos ({cols}) VALUES ({placeholders}) RETURNING id",
            list(filtered.values()),
        )
        new_id = cur.fetchone()[0]
        conn.commit()
        cur.close(); conn.close()
        return {"ok": True, "id": str(new_id)}
    except Exception as e:
        raise HTTPException(500, f"Error creando equipo: {str(e)[:200]}")


@app.delete("/admin/equipos/{equipo_id}")
async def eliminar_equipo(equipo_id: str):
    """Desactiva un equipo del catalogo (soft delete)."""
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("UPDATE equipos SET activo = FALSE WHERE id = %s", (equipo_id,))
        conn.commit()
        deleted = cur.rowcount
        cur.close(); conn.close()
        if deleted == 0:
            raise HTTPException(404, "Equipo no encontrado")
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Error: {str(e)[:200]}")


@app.post("/feedback")
async def recibir_feedback(data: dict):
    """
    Recibe feedback del usuario, lo envía a la IA para aplicar cambios,
    y regenera el ZIP completo con la propuesta modificada.
    """
    mensaje = data.get("mensaje", "").strip()
    if not mensaje:
        raise HTTPException(status_code=400, detail="Mensaje vacio")

    if not _ultimo_contexto.get("formulario"):
        raise HTTPException(status_code=400, detail="No hay propuesta previa. Genera una primero.")

    print(f"\n[FEEDBACK] Solicitud: {mensaje[:150]}")

    proyecto_nombre = ""
    try:
        proyecto_nombre = _ultimo_contexto.get("propuesta", {})
        proyecto_nombre = getattr(proyecto_nombre, "nombre_proyecto", "") or ""
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO feedback (mensaje, proyecto) VALUES (%s, %s)",
            (mensaje, proyecto_nombre),
        )
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        print(f"[FEEDBACK] WARN: No se pudo guardar en BD: {e}")

    formulario = _ultimo_contexto["formulario"]
    propuesta_anterior = _ultimo_contexto["propuesta"]
    plano_dxf = _ultimo_contexto.get("plano_dxf")
    descuento_general = _ultimo_contexto.get("descuento_general", 0.0) or 0.0
    descuentos_individuales = dict(_ultimo_contexto.get("descuentos_individuales") or {})
    sin_descuento_list = list(_ultimo_contexto.get("sin_descuento") or [])

    # Detectar y aplicar cambios de descuento desde el mensaje (no requiere LLM)
    import re as _re
    mensaje_residual = mensaje
    descuento_cambiado = False

    m_gen = _re.search(
        r"descuento\s+general[^0-9%]*([0-9]+(?:[.,][0-9]+)?)\s*%?",
        mensaje, _re.IGNORECASE,
    )
    if m_gen:
        nuevo_dto = float(m_gen.group(1).replace(",", "."))
        descuento_general = max(0.0, min(100.0, nuevo_dto))
        descuento_cambiado = True
        mensaje_residual = _re.sub(
            r"(cambia(?:r)?|pon(?:er|le)?|aplica(?:r)?)?\s*el?\s*descuento\s+general[^.,;]*[.,;]?",
            "", mensaje_residual, flags=_re.IGNORECASE,
        ).strip()
        print(f"[FEEDBACK] Descuento general -> {descuento_general}%")

    # Si tras quitar el cambio de descuento no queda nada relevante, saltar al LLM
    solo_descuento = descuento_cambiado and len(mensaje_residual) < 8

    try:
        from features.propuestas.llm import invocar_llm_con_rotacion, PropuestaEquipos
        from langchain_core.messages import SystemMessage, HumanMessage

        if solo_descuento:
            print("[FEEDBACK] Solo cambio de descuento, omitiendo LLM y reusando propuesta actual")
            nueva_propuesta = propuesta_anterior
        else:
            equipos_prev = []
            for zona_name, zona_list in [
                ("coccion", propuesta_anterior.zona_coccion),
                ("frio", propuesta_anterior.zona_frio),
                ("lavado", propuesta_anterior.zona_lavado),
                ("horno", propuesta_anterior.zona_horno),
            ]:
                for eq in zona_list:
                    equipos_prev.append(f"  - {eq.tipo} x{eq.cantidad} ({zona_name})")

            propuesta_json = propuesta_anterior.model_dump_json(indent=2)

            messages = [
                SystemMessage(content=(
                    "Eres un ingeniero de cocinas industriales RepagasConcept. "
                    "El usuario ya recibio una propuesta de equipamiento y quiere hacer cambios. "
                    "Tu tarea: aplicar EXACTAMENTE los cambios que pide y devolver la propuesta completa modificada. "
                    "Mantener todo lo que no se pida cambiar. Responde con el JSON estructurado PropuestaEquipos. "
                    "NOTA: los descuentos comerciales (general, individuales, exclusiones) NO forman parte de la "
                    "propuesta de equipamiento; ignoralos si el usuario los menciona."
                )),
                HumanMessage(content=(
                    f"PROPUESTA ACTUAL:\n{propuesta_json}\n\n"
                    f"CAMBIO SOLICITADO POR EL USUARIO:\n{mensaje_residual or mensaje}\n\n"
                    f"Aplica el cambio y devuelve la PropuestaEquipos completa modificada."
                )),
            ]

            print("[FEEDBACK] Enviando a la IA para aplicar cambios...")
            nueva_propuesta = invocar_llm_con_rotacion(messages, structured_cls=PropuestaEquipos)

            if not nueva_propuesta:
                raise HTTPException(500, "La IA no pudo procesar el cambio. Intenta reformular la solicitud.")

        print(f"[FEEDBACK] IA respondio. Regenerando propuesta...")

        prefs = (formulario.necesidades_equipamiento.preferencias_colocacion or "").lower()
        if "serie 900" in prefs or "fondo 900" in prefs:
            serie = "900"
        elif "serie 750" in prefs or "fondo 750" in prefs:
            serie = "750"
        else:
            serie = "900" if formulario.proyecto.comensales > 100 else "750"

        equipos_resueltos = resolver_equipos(nueva_propuesta, serie_pref=serie)

        _ultimo_contexto["propuesta"] = nueva_propuesta
        _ultimo_contexto["equipos_resueltos"] = equipos_resueltos
        _ultimo_contexto["descuento_general"] = descuento_general
        _ultimo_contexto["descuentos_individuales"] = descuentos_individuales
        _ultimo_contexto["sin_descuento"] = sin_descuento_list

        tmp_dir = os.path.join(UPLOAD_DIR, uuid.uuid4().hex)
        os.makedirs(tmp_dir, exist_ok=True)

        try:
            layout_tipo = getattr(nueva_propuesta, "layout", "L")
            dxf_path = os.path.join(tmp_dir, "propuesta.dxf")
            if plano_dxf:
                dxf_path, _plano_usado = generar_plano_integrado(
                    equipos_resueltos, plano_dxf, filepath=dxf_path, layout_tipo=layout_tipo
                )
            else:
                dxf_path = generar_plano(equipos_resueltos, filepath=dxf_path, layout_tipo=layout_tipo)
            png_path = dxf_path.replace(".dxf", ".png")

            total_pvp = sum((e.pvp_eur or 0) * e.cantidad for e in equipos_resueltos)
            resultado = {
                "proyecto": nueva_propuesta.nombre_proyecto,
                "layout": nueva_propuesta.layout,
                "equipos": [
                    {"modelo": e.modelo, "tipo": e.tipo, "ancho_mm": e.ancho_mm,
                     "fondo_mm": e.fondo_mm, "pvp_eur": e.pvp_eur, "cantidad": e.cantidad,
                     "zona": e.zona, "serie": e.serie}
                    for e in equipos_resueltos
                ],
                "total_equipos": len(equipos_resueltos),
                "total_pvp_eur": round(total_pvp, 2),
                "notas_llm": nueva_propuesta.notas,
                "feedback_aplicado": mensaje,
            }

            nombre_proy = nueva_propuesta.nombre_proyecto or ""
            formulario_dict = formulario.model_dump()
            from features.planos.posicionar import EquipoPosicionado
            equipos_pos = [
                EquipoPosicionado(
                    modelo=e.modelo, tipo=e.tipo, zona=e.zona,
                    ancho_mm=e.ancho_mm, fondo_mm=e.fondo_mm, alto_mm=e.alto_mm,
                    pvp_eur=e.pvp_eur, serie=e.serie, cantidad=1,
                    x=0, y=0, rotation=0, corners=None, wall_side="north",
                )
                for e in equipos_resueltos
                for _ in range(e.cantidad)
            ]

            pdf_prospeccion = os.path.join(tmp_dir, "prospeccion.pdf")
            try: generar_pdf_prospeccion(formulario_dict, nombre_proy, pdf_prospeccion)
            except Exception: pdf_prospeccion = None

            pdf_presupuesto = os.path.join(tmp_dir, "presupuesto.pdf")
            try:
                generar_pdf_presupuesto(
                    equipos_pos, nombre_proy, formulario_dict, pdf_presupuesto,
                    descuento_general=descuento_general,
                    descuentos_individuales=descuentos_individuales,
                    sin_descuento=sin_descuento_list,
                )
            except Exception:
                pdf_presupuesto = None

            # Tabla de equipos en DXF/DWG
            tabla_dxf_path = os.path.join(tmp_dir, "tabla_equipos.dxf")
            tabla_dwg_path = None
            try:
                from features.documentos.tabla import generar_tabla_equipos_dxf
                generar_tabla_equipos_dxf(equipos_resueltos, nombre_proy, tabla_dxf_path)
            except Exception as e:
                print(f"[FEEDBACK] WARN: tabla DXF fallo: {e}")
                tabla_dxf_path = None
            if tabla_dxf_path and os.path.isfile(tabla_dxf_path):
                try:
                    from features.planos.conversion import dxf_a_dwg as _dxf_a_dwg
                    tabla_dwg_path = _dxf_a_dwg(tabla_dxf_path, output_dir=tmp_dir)
                except Exception as e:
                    print(f"[FEEDBACK] WARN: tabla DWG fallo: {e}")

            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                zf.writestr("resultado.json", json.dumps(resultado, ensure_ascii=False, indent=2))
                if os.path.isfile(dxf_path):
                    zf.write(dxf_path, "propuesta.dxf")
                if os.path.isfile(png_path):
                    zf.write(png_path, "propuesta.png")
                if pdf_prospeccion and os.path.isfile(pdf_prospeccion):
                    zf.write(pdf_prospeccion, "prospeccion.pdf")
                if pdf_presupuesto and os.path.isfile(pdf_presupuesto):
                    zf.write(pdf_presupuesto, "presupuesto.pdf")
                if tabla_dwg_path and os.path.isfile(tabla_dwg_path):
                    zf.write(tabla_dwg_path, "tabla_equipos.dwg")
                if tabla_dxf_path and os.path.isfile(tabla_dxf_path):
                    zf.write(tabla_dxf_path, "tabla_equipos.dxf")

            zip_buffer.seek(0)
            print(f"[FEEDBACK] Propuesta regenerada con cambios aplicados")

            return StreamingResponse(
                zip_buffer,
                media_type="application/zip",
                headers={"Content-Disposition": "attachment; filename=propuesta_modificada.zip"},
            )
        finally:
            shutil.rmtree(tmp_dir, ignore_errors=True)

    except HTTPException:
        raise
    except Exception as e:
        code, msg = _humanizar_error(e, contexto="aplicacion de feedback sobre la propuesta")
        raise HTTPException(code, msg)


@app.get("/")
def root():
    return {"status": "ok", "mensaje": "Repagas Generador de Cocinas - API activa"}


@app.get("/catalogo")
def obtener_catalogo():
    """
    Devuelve el catálogo de equipos agrupado por zona (coccion, refrigeracion, lavado, horno).
    Cada equipo incluye: modelo, tipo, ancho_mm, fondo_mm, alto_mm, pvp_eur, serie.
    """
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            SELECT e.modelo, e.tipo, e.ancho_mm, e.fondo_mm, e.alto_mm,
                   e.pvp_eur, s.nombre as serie, e.alimentacion
            FROM equipos e
            LEFT JOIN series s ON e.serie_id = s.id
            WHERE e.activo = TRUE
              AND e.ancho_mm IS NOT NULL
              AND e.fondo_mm IS NOT NULL
            ORDER BY e.tipo, e.ancho_mm, e.modelo
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        raise HTTPException(500, f"Error consultando base de datos: {e}")

    catalogo = {"coccion": [], "refrigeracion": [], "lavado": [], "horno": []}
    for row in rows:
        modelo, tipo, ancho, fondo, alto, pvp, serie, alim = row
        zona = TIPO_TO_ZONA.get(tipo, "coccion")
        catalogo[zona].append({
            "modelo": modelo,
            "tipo": tipo,
            "ancho_mm": ancho,
            "fondo_mm": fondo,
            "alto_mm": alto,
            "pvp_eur": float(pvp) if pvp else None,
            "serie": serie or "",
            "alimentacion": alim or "",
        })

    return catalogo


@app.post("/generar")
async def generar_cocina(request: Request):
    """Recibe JSON con FormularioCliente y devuelve ZIP con DWG + PDFs + resultado.json."""
    try:
        data = await request.json()
        incluir_dxf = bool(data.pop("_incluir_dxf", False))
        descuento = float(data.pop("_descuento_general", 0) or 0)
        descuentos_ind = data.pop("_descuentos_individuales", None) or {}
        sin_dto = data.pop("_sin_descuento", None) or []
        openrouter_model = data.pop("_openrouter_model", None) or None
        formulario = FormularioCliente(**data)
    except Exception as e:
        code, msg = _humanizar_error(e, contexto="validacion del formulario")
        raise HTTPException(code, msg)
    try:
        return _ejecutar_pipeline(formulario, incluir_dxf=incluir_dxf,
                                   descuento_general=descuento,
                                   descuentos_individuales=descuentos_ind,
                                   sin_descuento=sin_dto,
                                   openrouter_model=openrouter_model)
    except HTTPException:
        raise
    except Exception as e:
        code, msg = _humanizar_error(e, contexto="generacion de la propuesta")
        raise HTTPException(code, msg)


@app.post("/generar-con-plano")
async def generar_con_plano(
    formulario_json: str = Form(..., description="JSON string con datos del FormularioCliente"),
    archivo: UploadFile = File(..., description="Archivo .dwg o .dxf del plano del cliente"),
    incluir_dxf: bool = Form(False, description="Incluir archivo DXF ademas del DWG"),
    descuento_general: float = Form(0.0, description="Descuento en % aplicado a todos los equipos"),
    descuentos_individuales: str = Form("{}", description="JSON con descuentos por modelo"),
    sin_descuento: str = Form("[]", description="JSON list de modelos sin descuento"),
    openrouter_model: str = Form("", description="Modelo OpenRouter a usar (vacio = default)"),
):
    """
    Igual que /generar pero recibe el formulario como form-data junto al DWG/DXF del cliente.
    """
    try:
        data = json.loads(formulario_json)
        for k in ("_incluir_dxf", "_descuento_general", "_descuentos_individuales", "_sin_descuento", "_openrouter_model"):
            data.pop(k, None)
        formulario = FormularioCliente(**data)
    except json.JSONDecodeError as e:
        raise HTTPException(400, f"JSON invalido en formulario_json: {e}")
    except Exception as e:
        raise HTTPException(400, f"Error en datos del formulario: {e}")

    try:
        descuentos_ind = json.loads(descuentos_individuales) if descuentos_individuales else {}
    except Exception:
        descuentos_ind = {}
    try:
        sin_dto = json.loads(sin_descuento) if sin_descuento else []
    except Exception:
        sin_dto = []

    try:
        ruta_archivo = _guardar_upload(archivo)
        plano_dxf = _convertir_si_dwg(ruta_archivo)
        print(f"[WEBHOOK] Plano del cliente: {plano_dxf}")
    except HTTPException:
        raise
    except Exception as e:
        code, msg = _humanizar_error(e, contexto="procesado del plano del cliente")
        raise HTTPException(code, msg)

    try:
        return _ejecutar_pipeline(formulario, plano_dxf=plano_dxf,
                                   incluir_dxf=incluir_dxf,
                                   descuento_general=descuento_general,
                                   descuentos_individuales=descuentos_ind,
                                   sin_descuento=sin_dto,
                                   openrouter_model=openrouter_model or None)
    except HTTPException:
        raise
    except Exception as e:
        code, msg = _humanizar_error(e, contexto="generacion de la propuesta con plano del cliente")
        raise HTTPException(code, msg)


# ─── Arranque directo ────────────────────────────────────

if __name__ == "__main__":
    import uvicorn
    print("Iniciando servidor webhook Repagas...")
    print("Docs interactivos en: http://localhost:8000/docs")
    uvicorn.run(app, host="0.0.0.0", port=8000)
